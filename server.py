#!/usr/bin/env python3
"""FirstPass v2 — HTTP API Server (port 8765)"""
import json, os, sys, traceback, threading
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import model as rft_model

PORT = 8765
_arts  = None
_lock  = threading.Lock()

def get_arts():
    global _arts
    with _lock:
        if _arts is None:
            _arts = (rft_model.load_artifacts() if os.path.exists(rft_model.MODEL_PATH)
                     else rft_model.train_and_save())
    return _arts

class Handler(BaseHTTPRequestHandler):
    def log_message(self, *a): pass   # silence access log

    def _cors(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET,POST,OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')

    def send_json(self, data, status=200):
        body = json.dumps(data, default=str).encode('utf-8')
        self.send_response(status)
        self.send_header('Content-Type',   'application/json; charset=utf-8')
        self.send_header('Content-Length', len(body))
        self._cors(); self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self):
        self.send_response(200); self._cors(); self.end_headers()

    def do_GET(self):
        path = urlparse(self.path).path

        # ── Serve UI ──────────────────────────────────────────────────────────
        if path == '/':
            html_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'index.html')
            body = open(html_path, 'rb').read()
            self.send_response(200)
            self.send_header('Content-Type',   'text/html; charset=utf-8')
            self.send_header('Content-Length', len(body))
            self.end_headers(); self.wfile.write(body); return

        # ── Dashboard data ────────────────────────────────────────────────────
        elif path == '/api/dashboard':
            arts = get_arts()
            # Strip non-serialisable ML objects; send everything else
            payload = {k: v for k, v in arts.items()
                       if k not in ('model', 'encoders')}
            self.send_json(payload); return

        # ── Retrain ───────────────────────────────────────────────────────────
        elif path == '/api/retrain':
            try:
                global _arts
                arts = rft_model.train_and_save()
                with _lock: _arts = arts
                self.send_json({'status': 'success', 'METRICS': arts['METRICS']}); return
            except Exception as e:
                self.send_json({'status': 'error', 'message': str(e),
                                'trace': traceback.format_exc()}, 500); return

        self.send_json({'error': 'Not found'}, 404)

    def do_POST(self):
        path   = urlparse(self.path).path
        length = int(self.headers.get('Content-Length', 0))
        raw    = self.rfile.read(length)
        try:    data = json.loads(raw) if raw else {}
        except: data = {}

        # ── Predict ───────────────────────────────────────────────────────────
        if path == '/api/predict':
            try:
                result = rft_model.predict_single(data, get_arts())
                self.send_json(result); return
            except Exception as e:
                self.send_json({'error': str(e), 'trace': traceback.format_exc()}, 500); return

        # ── Append data row (editor) ──────────────────────────────────────────
        elif path == '/api/data/add':
            try:
                from openpyxl import load_workbook
                wb = load_workbook(rft_model.DATA_PATH)
                ws = wb['ML_Training_Data']
                hdrs = [ws.cell(row=2, column=c).value
                        for c in range(1, ws.max_column + 1)]
                last = ws.max_row + 1
                for ci, col in enumerate(hdrs, 1):
                    ws.cell(row=last, column=ci, value=data.get(col, ''))
                wb.save(rft_model.DATA_PATH)
                self.send_json({'status': 'success', 'row': last}); return
            except Exception as e:
                self.send_json({'error': str(e)}, 500); return

        self.send_json({'error': 'Not found'}, 404)


def run():
    print('Pre-loading model artifacts…')
    get_arts()
    print(f'Server listening on http://localhost:{PORT}')
    HTTPServer(('localhost', PORT), Handler).serve_forever()

if __name__ == '__main__':
    run()
